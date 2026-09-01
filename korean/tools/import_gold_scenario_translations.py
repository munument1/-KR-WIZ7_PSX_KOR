#!/usr/bin/env python3
"""Import Wizardry 7 Gold Scenario translations into PS1 scenario TSVs.

The importer validates Gold `source_text` against DOS-derived source work TSVs by ID.
Item Gold names may be a prefix of the full DOS name because the Gold table uses a
shorter fixed field; monster names must match exactly.  An optional overrides TSV
keeps manual QA fixes reproducible.
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

ITEM_LIMIT = 21
MONSTER_LIMIT = 15
MON_VARIANTS = {
    "singular": "specific_singular",
    "plural": "specific_plural",
    "generic_singular": "generic_singular",
    "generic_plural": "generic_plural",
}
NORMALIZE = {
    "\u2018": "'",
    "\u2019": "'",
    "\u201c": '"',
    "\u201d": '"',
    "\u00a0": " ",
    "\u00d7": "x",
}


def read_table(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise ValueError(".xlsx input requires openpyxl; export Scenario as TSV/CSV instead") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        if "Scenario" not in wb.sheetnames:
            raise ValueError("xlsx has no Scenario sheet")
        ws = wb["Scenario"]
        rows = ws.iter_rows(values_only=True)
        header = [str(v or "") for v in next(rows)]
        return [
            {header[i]: str(v or "") for i, v in enumerate(row[: len(header)])}
            for row in rows
            if any(v is not None for v in row)
        ]

    delimiter = "," if suffix == ".csv" else "\t"
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f, delimiter=delimiter))


def encoded_len(text: str) -> int:
    total = 0
    for ch in text:
        for c in NORMALIZE.get(ch, ch):
            total += 1 if ord(c) <= 0x7F else 2
    return total


def load_overrides(path: Path | None) -> dict[tuple[str, int, str], str]:
    if path is None:
        return {}
    out: dict[tuple[str, int, str], str] = {}
    for row in read_table(path):
        target = row.get("target", "").strip()
        record_id = int(row.get("id", "-1"))
        field = row.get("field", "").strip()
        translation = row.get("translation", "").strip()
        if target not in {"item", "monster"}:
            raise ValueError(f"bad override target: {target!r}")
        key = (target, record_id, field)
        if key in out:
            raise ValueError(f"duplicate override: {key}")
        out[key] = translation
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--gold-scenario", type=Path, required=True,
                    help="Gold Scenario sheet exported as TSV/CSV, or the full .xlsx workbook")
    ap.add_argument("--items-source-work", type=Path, required=True)
    ap.add_argument("--monsters-source-work", type=Path, required=True)
    ap.add_argument("--overrides", type=Path)
    ap.add_argument("--out-items", type=Path, required=True)
    ap.add_argument("--out-monsters", type=Path, required=True)
    args = ap.parse_args()

    gold_rows = read_table(args.gold_scenario)
    gold_items: dict[int, tuple[str, str]] = {}
    gold_monsters: dict[tuple[int, str], tuple[str, str]] = {}
    for row in gold_rows:
        category = row.get("category", "").strip()
        if category not in {"item", "monster"}:
            continue
        rid = int(row["record_index"])
        variant = row.get("variant", "").strip()
        pair = (row.get("source_text", "").strip(), row.get("translation", "").strip())
        if category == "item":
            gold_items[rid] = pair
        else:
            gold_monsters[(rid, variant)] = pair

    item_src = read_table(args.items_source_work)
    mon_src = read_table(args.monsters_source_work)
    overrides = load_overrides(args.overrides)

    prefix_trunc = 0
    item_bad: list[str] = []
    monster_bad: list[str] = []
    overlength: list[str] = []

    args.out_items.parent.mkdir(parents=True, exist_ok=True)
    with args.out_items.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", lineterminator="\n")
        w.writerow(["id", "ko_name"])
        for row in item_src:
            rid = int(row["id"])
            full = row.get("source_en", "")
            gold_src, translation = gold_items.get(rid, ("", ""))
            if gold_src != full:
                if gold_src and full.startswith(gold_src):
                    prefix_trunc += 1
                elif gold_src or full:
                    item_bad.append(f"item {rid}: DOS={full!r} Gold={gold_src!r}")
            translation = overrides.get(("item", rid, "name"), translation)
            size = encoded_len(translation)
            if translation and size > ITEM_LIMIT:
                overlength.append(f"item {rid}: {size}>{ITEM_LIMIT} {translation!r}")
            w.writerow([rid, translation])

    with args.out_monsters.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", lineterminator="\n")
        w.writerow(["id", "ko_specific_singular", "ko_specific_plural",
                    "ko_generic_singular", "ko_generic_plural"])
        for row in mon_src:
            rid = int(row["id"])
            out = [rid]
            for gold_variant, src_field in MON_VARIANTS.items():
                expected = row.get("en_" + src_field, "")
                gold_src, translation = gold_monsters.get((rid, gold_variant), ("", ""))
                if gold_src != expected:
                    monster_bad.append(
                        f"monster {rid} {gold_variant}: DOS={expected!r} Gold={gold_src!r}"
                    )
                translation = overrides.get(("monster", rid, gold_variant), translation)
                size = encoded_len(translation)
                if translation and size > MONSTER_LIMIT:
                    overlength.append(
                        f"monster {rid} {gold_variant}: {size}>{MONSTER_LIMIT} {translation!r}"
                    )
                out.append(translation)
            w.writerow(out)

    print(f"items={len(item_src)} monsters={len(mon_src)}")
    print(f"item_prefix_truncations={prefix_trunc}")
    print(f"item_source_mismatches={len(item_bad)}")
    print(f"monster_source_mismatches={len(monster_bad)}")
    print(f"overlength={len(overlength)}")

    for line in item_bad + monster_bad + overlength:
        print(line, file=sys.stderr)
    return 2 if item_bad or monster_bad or overlength else 0


if __name__ == "__main__":
    raise SystemExit(main())
